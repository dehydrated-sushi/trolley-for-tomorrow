"""Build combined ingredient nutrition, CPI, and expiry reference data.

Inputs default to the local files the project team has been using:
- /Users/arsh/Desktop/data/AFCD Release 3 - Nutrient profiles.xlsx
- /Users/arsh/Desktop/data/CPI-Australia/640103.xlsx
- data/processed/foodkeeper.json
- backend/data/processed/known_ingredients.csv

Outputs default to backend/data/processed so the backend can load them without
another copy step:
- afcd_nutrition_clean.csv
- cpi_food_indexes.csv
- expiry_reference.csv
- ingredient_expiry_matches.csv
- food_reference.csv

Run:
    python data/pipelines/build_food_reference.py
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DESKTOP_DATA = Path("/Users/arsh/Desktop/data")
DEFAULT_AFCD = DEFAULT_DESKTOP_DATA / "AFCD Release 3 - Nutrient profiles.xlsx"
DEFAULT_CPI = DEFAULT_DESKTOP_DATA / "CPI-Australia" / "640103.xlsx"
DEFAULT_FOODKEEPER = REPO_ROOT / "data" / "processed" / "foodkeeper.json"
DEFAULT_KNOWN = REPO_ROOT / "backend" / "data" / "processed" / "known_ingredients.csv"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "backend" / "data" / "processed"

TOKEN_STOPWORDS = {
    "added", "and", "black", "bottle", "bottled", "bought", "can", "canned",
    "chopped", "commercial", "cooked", "diced", "dried", "dry", "extra",
    "fat", "flavored", "flavour", "free", "fresh", "frozen", "green",
    "homemade", "in", "jar", "large", "lean", "low", "medium", "minced",
    "no", "of", "organic", "package", "pack", "pieces", "piece", "pkg",
    "plain", "raw", "reduced", "red", "ripe", "salted", "sliced", "slice",
    "small", "store", "style", "sweet", "sweetened", "the", "to",
    "uncooked", "unsalted", "unsweetened", "white", "whole", "with",
    "yellow",
}

GENERIC_SINGLE_TOKEN_ALIASES = {
    "blend", "boneless", "bottled", "butter", "dinner", "drink", "flavor",
    "food", "french", "meal", "mix", "nectar", "oil", "paste", "powder",
    "ready", "sauce", "seed", "seeds", "steak", "steaks", "sweet", "syrup",
    "uncooked", "water", "wild",
}

DISQUALIFYING_PRODUCT_TOKENS = {"stuffed", "vegan"}

NUTRITION_COLUMNS = [
    "food_key",
    "classification",
    "derivation",
    "food_name",
    "canonical_name",
    "basis",
    "energy_kj",
    "energy_kcal",
    "protein_g",
    "fat_total_g",
    "carbohydrates_g",
    "sugars_g",
    "fibre_g",
    "sodium_mg",
    "calcium_mg",
    "iron_mg",
    "source",
]

CPI_COLUMNS = [
    "period",
    "series_id",
    "measure",
    "cpi_category",
    "geography",
    "unit",
    "series_type",
    "data_type",
    "value",
    "source_file",
]

EXPIRY_COLUMNS = [
    "foodkeeper_product_id",
    "foodkeeper_name",
    "foodkeeper_subtitle",
    "foodkeeper_keywords",
    "foodkeeper_category",
    "foodkeeper_subcategory",
    "pantry_min_days",
    "pantry_max_days",
    "pantry_metric",
    "pantry_tips",
    "refrigerate_min_days",
    "refrigerate_max_days",
    "refrigerate_metric",
    "refrigerate_tips",
    "freeze_min_days",
    "freeze_max_days",
    "freeze_metric",
    "freeze_tips",
    "source",
]

EXPIRY_MATCH_COLUMNS = [
    "ingredient_name",
    "foodkeeper_product_id",
    "foodkeeper_name",
    "foodkeeper_subtitle",
    "match_score",
    "match_source",
    "matched_alias",
]

FOOD_REFERENCE_COLUMNS = [
    "ingredient_name",
    "canonical_name",
    "afcd_food_key",
    "afcd_food_name",
    "afcd_match_score",
    "nutrition_basis",
    "energy_kj",
    "energy_kcal",
    "protein_g",
    "fat_total_g",
    "carbohydrates_g",
    "sugars_g",
    "fibre_g",
    "sodium_mg",
    "cpi_category",
    "latest_cpi_period",
    "latest_cpi_index",
    "cpi_monthly_pct_change",
    "cpi_annual_pct_change",
    "foodkeeper_product_id",
    "foodkeeper_name",
    "foodkeeper_subtitle",
    "expiry_match_score",
    "pantry_min_days",
    "pantry_max_days",
    "refrigerate_min_days",
    "refrigerate_max_days",
    "freeze_min_days",
    "freeze_max_days",
]


def repair_mojibake(value):
    if not isinstance(value, str) or not value:
        return value
    if "Ã" not in value and "â" not in value:
        return value
    try:
        repaired = value.encode("latin-1", errors="ignore").decode("utf-8", errors="replace")
    except Exception:
        return value
    before = value.count("Ã") + value.count("â")
    after = repaired.count("Ã") + repaired.count("â")
    return repaired if after < before else value


def clean_text(value):
    if value is None:
        return ""
    return repair_mojibake(str(value)).strip()


def normalise_text(value):
    text = clean_text(value).lower().replace("&", " and ")
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", errors="ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def singular_token(token):
    if len(token) > 4 and token.endswith("ies"):
        return token[:-3] + "y"
    if len(token) > 4 and token.endswith(("ches", "shes", "xes", "zes")):
        return token[:-2]
    if len(token) > 3 and token.endswith("s"):
        return token[:-1]
    return token


def match_key(value):
    tokens = []
    for token in normalise_text(value).split():
        if len(token) <= 1 or token in TOKEN_STOPWORDS:
            continue
        tokens.append(singular_token(token))
    return " ".join(tokens)


def token_set(value):
    return set(match_key(value).split())


def number_or_blank(value):
    if value in (None, ""):
        return ""
    try:
        return float(value)
    except (TypeError, ValueError):
        return ""


def int_or_blank(value):
    if value in (None, ""):
        return ""
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return ""


def write_csv(path, rows, columns):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def read_known_ingredients(path):
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if "ingredient_name" not in (reader.fieldnames or []):
            raise ValueError(f"{path} must contain ingredient_name")
        names = [clean_text(row.get("ingredient_name")) for row in reader]

    seen = set()
    out = []
    for name in names:
        key = name.lower()
        if name and key not in seen:
            seen.add(key)
            out.append(name)
    return out


def clean_header(value):
    return re.sub(r"\s+", " ", clean_text(value)).strip().lower()


def find_header(headers, *needles):
    for index, header in enumerate(headers):
        header_clean = clean_header(header)
        if all(needle in header_clean for needle in needles):
            return index
    return None


def get_by_header(row, headers, *needles):
    index = find_header(headers, *needles)
    if index is None or index >= len(row):
        return ""
    return row[index]


def load_afcd_nutrition(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = []
    sheet_specs = [
        ("All solids & liquids per 100 g", "per_100_g"),
        ("Liquids only per 100 mL", "per_100_ml"),
    ]

    for sheet_name, basis in sheet_specs:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
        for sheet_row in ws.iter_rows(min_row=4, values_only=True):
            food_key = clean_text(get_by_header(sheet_row, headers, "public food key"))
            food_name = clean_text(get_by_header(sheet_row, headers, "food name"))
            if not food_key or not food_name:
                continue
            energy_kj = number_or_blank(get_by_header(sheet_row, headers, "energy with dietary fibre"))
            energy_kcal = round(energy_kj * 0.239006, 2) if energy_kj != "" else ""
            rows.append({
                "food_key": food_key,
                "classification": int_or_blank(get_by_header(sheet_row, headers, "classification")),
                "derivation": clean_text(get_by_header(sheet_row, headers, "derivation")),
                "food_name": food_name,
                "canonical_name": match_key(food_name),
                "basis": basis,
                "energy_kj": energy_kj,
                "energy_kcal": energy_kcal,
                "protein_g": number_or_blank(get_by_header(sheet_row, headers, "protein")),
                "fat_total_g": number_or_blank(get_by_header(sheet_row, headers, "fat, total")),
                "carbohydrates_g": number_or_blank(
                    get_by_header(sheet_row, headers, "available carbohydrate", "without sugar alcohols")
                ),
                "sugars_g": number_or_blank(get_by_header(sheet_row, headers, "total sugars")),
                "fibre_g": number_or_blank(get_by_header(sheet_row, headers, "total dietary fibre")),
                "sodium_mg": number_or_blank(get_by_header(sheet_row, headers, "sodium")),
                "calcium_mg": number_or_blank(get_by_header(sheet_row, headers, "calcium")),
                "iron_mg": number_or_blank(get_by_header(sheet_row, headers, "iron")),
                "source": "AFCD Release 3",
            })
    return rows


def parse_cpi_descriptor(value):
    parts = [clean_text(part) for part in clean_text(value).split(";")]
    parts = [part for part in parts if part]
    return {
        "measure": parts[0] if len(parts) > 0 else "",
        "cpi_category": parts[1] if len(parts) > 1 else "",
        "geography": parts[2] if len(parts) > 2 else "",
    }


def load_cpi_indexes(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Data1"]
    descriptor_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    unit_row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    series_type_row = [cell.value for cell in next(ws.iter_rows(min_row=3, max_row=3))]
    data_type_row = [cell.value for cell in next(ws.iter_rows(min_row=4, max_row=4))]
    series_id_row = [cell.value for cell in next(ws.iter_rows(min_row=10, max_row=10))]

    rows = []
    for col_index in range(1, len(descriptor_row)):
        descriptor = parse_cpi_descriptor(descriptor_row[col_index])
        if descriptor["measure"] != "Index Numbers":
            continue
        if descriptor["geography"] != "Australia":
            continue
        for sheet_row in ws.iter_rows(min_row=11, values_only=True):
            period = sheet_row[0]
            value = sheet_row[col_index] if col_index < len(sheet_row) else None
            if not isinstance(period, datetime) or value in (None, ""):
                continue
            rows.append({
                "period": period.date().isoformat(),
                "series_id": clean_text(series_id_row[col_index]),
                "measure": descriptor["measure"],
                "cpi_category": descriptor["cpi_category"],
                "geography": descriptor["geography"],
                "unit": clean_text(unit_row[col_index]),
                "series_type": clean_text(series_type_row[col_index]),
                "data_type": clean_text(data_type_row[col_index]),
                "value": number_or_blank(value),
                "source_file": path.name,
            })
    return rows


def latest_cpi_by_category(cpi_rows):
    by_category = defaultdict(list)
    for row in cpi_rows:
        by_category[row["cpi_category"]].append(row)

    latest = {}
    for category, rows in by_category.items():
        ordered = sorted(rows, key=lambda row: row["period"])
        last = ordered[-1]
        previous = ordered[-2] if len(ordered) >= 2 else None
        prior_year_period = last["period"][:4]
        annual_base = None
        try:
            annual_year = str(int(prior_year_period) - 1)
            annual_date = annual_year + last["period"][4:]
            annual_base = next((r for r in ordered if r["period"] == annual_date), None)
        except ValueError:
            annual_base = None

        def pct_change(base):
            if not base or base.get("value") in ("", 0):
                return ""
            return round(((last["value"] - base["value"]) / base["value"]) * 100, 2)

        latest[category] = {
            "latest_cpi_period": last["period"],
            "latest_cpi_index": last["value"],
            "cpi_monthly_pct_change": pct_change(previous),
            "cpi_annual_pct_change": pct_change(annual_base),
        }
    return latest


def flatten_foodkeeper_row(row):
    merged = {}
    for cell in row:
        if isinstance(cell, dict):
            merged.update(cell)
    return merged


def foodkeeper_int(value):
    if value in (None, ""):
        return ""
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return ""


def days_from_metric(value, metric):
    value = foodkeeper_int(value)
    metric = clean_text(metric).lower()
    if value == "":
        return ""
    if metric.startswith("day"):
        return value
    if metric.startswith("week"):
        return value * 7
    if metric.startswith("month"):
        return value * 30
    if metric.startswith("year"):
        return value * 365
    return ""


def load_foodkeeper_expiry(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    sheets = {sheet["name"]: sheet.get("data", []) for sheet in data.get("sheets", [])}

    categories = {}
    for row in sheets.get("Category", []):
        r = flatten_foodkeeper_row(row)
        category_id = foodkeeper_int(r.get("ID"))
        if category_id != "":
            categories[category_id] = {
                "category": clean_text(r.get("Category_Name")),
                "subcategory": clean_text(r.get("Subcategory_Name")),
            }

    rows = []
    raw_products = []
    for row in sheets.get("Product", []):
        r = flatten_foodkeeper_row(row)
        product_id = foodkeeper_int(r.get("ID"))
        name = clean_text(r.get("Name"))
        if product_id == "" or not name:
            continue
        category = categories.get(foodkeeper_int(r.get("Category_ID")), {})
        item = {
            "foodkeeper_product_id": product_id,
            "foodkeeper_name": name,
            "foodkeeper_subtitle": clean_text(r.get("Name_subtitle")),
            "foodkeeper_keywords": clean_text(r.get("Keywords")),
            "foodkeeper_category": category.get("category", ""),
            "foodkeeper_subcategory": category.get("subcategory", ""),
            "pantry_min_days": days_from_metric(r.get("DOP_Pantry_Min") or r.get("Pantry_Min"), r.get("DOP_Pantry_Metric") or r.get("Pantry_Metric")),
            "pantry_max_days": days_from_metric(r.get("DOP_Pantry_Max") or r.get("Pantry_Max"), r.get("DOP_Pantry_Metric") or r.get("Pantry_Metric")),
            "pantry_metric": clean_text(r.get("DOP_Pantry_Metric") or r.get("Pantry_Metric")),
            "pantry_tips": clean_text(r.get("DOP_Pantry_tips") or r.get("Pantry_tips")),
            "refrigerate_min_days": days_from_metric(r.get("DOP_Refrigerate_Min") or r.get("Refrigerate_Min"), r.get("DOP_Refrigerate_Metric") or r.get("Refrigerate_Metric")),
            "refrigerate_max_days": days_from_metric(r.get("DOP_Refrigerate_Max") or r.get("Refrigerate_Max"), r.get("DOP_Refrigerate_Metric") or r.get("Refrigerate_Metric")),
            "refrigerate_metric": clean_text(r.get("DOP_Refrigerate_Metric") or r.get("Refrigerate_Metric")),
            "refrigerate_tips": clean_text(r.get("DOP_Refrigerate_tips") or r.get("Refrigerate_tips")),
            "freeze_min_days": days_from_metric(r.get("DOP_Freeze_Min") or r.get("Freeze_Min"), r.get("DOP_Freeze_Metric") or r.get("Freeze_Metric")),
            "freeze_max_days": days_from_metric(r.get("DOP_Freeze_Max") or r.get("Freeze_Max"), r.get("DOP_Freeze_Metric") or r.get("Freeze_Metric")),
            "freeze_metric": clean_text(r.get("DOP_Freeze_Metric") or r.get("Freeze_Metric")),
            "freeze_tips": clean_text(r.get("DOP_Freeze_Tips") or r.get("Freeze_Tips")),
            "source": "USDA FSIS FoodKeeper Data v128",
        }
        rows.append(item)
        raw_products.append({
            "id": product_id,
            "name": item["foodkeeper_name"],
            "subtitle": item["foodkeeper_subtitle"],
            "keywords": item["foodkeeper_keywords"],
        })
    return rows, raw_products


def build_expiry_aliases(products):
    aliases = []
    seen = set()
    for product in products:
        candidates = [
            ("name", product["name"]),
            ("subtitle", product.get("subtitle")),
            ("combined", f"{product['name']} {product.get('subtitle') or ''}".strip()),
        ]
        candidates.extend(("keyword", keyword) for keyword in (product.get("keywords") or "").split(","))
        product_tokens = token_set(product["name"])
        for source, value in candidates:
            alias = normalise_text(value)
            key = match_key(value)
            if not key:
                continue
            dedupe = (product["id"], source, key)
            if dedupe in seen:
                continue
            seen.add(dedupe)
            aliases.append({
                "alias": alias,
                "key": key,
                "tokens": set(key.split()),
                "product": product,
                "product_tokens": product_tokens,
                "source": source,
            })
    return aliases


def score_expiry_match(ingredient_name, aliases):
    ingredient_key = match_key(ingredient_name)
    ingredient_tokens = set(ingredient_key.split())
    if not ingredient_key:
        return None
    best = None

    def consider(score, product, source, alias):
        nonlocal best
        if best is None or score > best["match_score"]:
            best = {
                "ingredient_name": ingredient_name,
                "foodkeeper_product_id": product["id"],
                "foodkeeper_name": product["name"],
                "foodkeeper_subtitle": product.get("subtitle", ""),
                "match_score": round(score, 3),
                "match_source": source,
                "matched_alias": alias,
            }

    padded_ingredient = f" {ingredient_key} "
    for alias in aliases:
        if (alias["product_tokens"] & DISQUALIFYING_PRODUCT_TOKENS) - ingredient_tokens:
            continue
        alias_key = alias["key"]
        alias_tokens = alias["tokens"]
        token_count = len(alias_tokens)
        if ingredient_key == alias_key:
            if alias["source"] == "name":
                consider(1.0, alias["product"], "exact_name", alias["alias"])
            elif alias["source"] == "combined":
                consider(0.97, alias["product"], "exact_combined", alias["alias"])
            elif alias["source"] == "keyword" and (
                token_count > 1 or alias_key not in GENERIC_SINGLE_TOKEN_ALIASES
            ):
                consider(0.93, alias["product"], "exact_keyword", alias["alias"])
            continue
        if token_count == 1:
            continue
        if f" {alias_key} " in padded_ingredient:
            consider(0.88, alias["product"], "foodkeeper_phrase_in_ingredient", alias["alias"])
            continue
        overlap = len(ingredient_tokens & alias_tokens)
        if overlap < 2:
            continue
        score = (0.75 * (overlap / len(ingredient_tokens))) + (0.25 * (overlap / len(alias_tokens)))
        if score >= 0.78:
            consider(score, alias["product"], "token_overlap", alias["alias"])
    return best if best and best["match_score"] >= 0.78 else None


def build_match_index(rows, name_column):
    index = defaultdict(set)
    token_cache = []
    for i, row in enumerate(rows):
        tokens = token_set(row[name_column])
        token_cache.append(tokens)
        for token in tokens:
            index[token].add(i)
    return index, token_cache


def best_token_match(name, rows, name_column, index, token_cache, threshold=0.55):
    query_tokens = token_set(name)
    if not query_tokens:
        return None, ""
    candidate_ids = set()
    for token in query_tokens:
        candidate_ids.update(index.get(token, set()))
    best = None
    for candidate_id in candidate_ids:
        row_tokens = token_cache[candidate_id]
        if not row_tokens:
            continue
        overlap = len(query_tokens & row_tokens)
        if overlap == 0:
            continue
        score = (0.75 * (overlap / len(query_tokens))) + (0.25 * (overlap / len(row_tokens)))
        if match_key(name) == match_key(rows[candidate_id][name_column]):
            score = 1.0
        if score >= threshold and (best is None or score > best[1]):
            best = (rows[candidate_id], round(score, 3))
    return best if best else (None, "")


def infer_cpi_category(name, available_categories):
    text = normalise_text(name)
    words = set(text.split())

    def has_keyword(keyword):
        keyword = normalise_text(keyword)
        if " " in keyword:
            return f" {keyword} " in f" {text} "
        return keyword in words

    rules = [
        ("Bread", ["bread", "bagel", "roll", "loaf"]),
        ("Cakes and biscuits", ["cake", "biscuit", "cookie", "muffin", "pastry"]),
        ("Breakfast cereals", ["cereal", "muesli", "granola"]),
        ("Beef and veal", ["beef", "veal"]),
        ("Pork", ["pork", "bacon", "ham"]),
        ("Lamb and goat", ["lamb", "goat"]),
        ("Poultry", ["chicken", "turkey", "duck"]),
        ("Fish and other seafood", ["fish", "salmon", "tuna", "seafood", "prawn", "shrimp", "crab"]),
        ("Milk", ["milk"]),
        ("Cheese", ["cheese", "cheddar", "mozzarella", "parmesan"]),
        ("Eggs", ["egg"]),
        ("Waters, soft drinks and juices", ["water", "juice", "soda", "soft drink", "cola"]),
        ("Fruit", ["apple", "banana", "orange", "berry", "berries", "pear", "mango", "grape"]),
        ("Vegetables", ["vegetable", "lettuce", "tomato", "potato", "onion", "carrot", "broccoli", "spinach"]),
        ("Coffee, tea and cocoa", ["coffee", "tea", "cocoa"]),
        ("Oils and fats", ["oil", "butter", "margarine", "fat"]),
        ("Sugar, jam, honey, chocolate and confectionery", ["sugar", "jam", "honey", "chocolate", "candy"]),
        ("Snacks and confectionery", ["chips", "cracker", "snack"]),
    ]
    for category, keywords in rules:
        if category in available_categories and any(has_keyword(keyword) for keyword in keywords):
            return category
    fallback = "Food and non-alcoholic beverages"
    return fallback if fallback in available_categories else ""


def build_food_reference(known, nutrition_rows, latest_cpi, expiry_rows, expiry_matches):
    nutrition_index, nutrition_token_cache = build_match_index(nutrition_rows, "food_name")
    expiry_by_ingredient = {row["ingredient_name"].lower(): row for row in expiry_matches}
    expiry_by_product = {row["foodkeeper_product_id"]: row for row in expiry_rows}
    available_cpi_categories = set(latest_cpi)

    rows = []
    for ingredient_name in known:
        nutrition, nutrition_score = best_token_match(
            ingredient_name,
            nutrition_rows,
            "food_name",
            nutrition_index,
            nutrition_token_cache,
        )
        cpi_category = infer_cpi_category(ingredient_name, available_cpi_categories)
        cpi = latest_cpi.get(cpi_category, {})
        expiry_match = expiry_by_ingredient.get(ingredient_name.lower(), {})
        expiry = expiry_by_product.get(expiry_match.get("foodkeeper_product_id"), {})
        row = {
            "ingredient_name": ingredient_name,
            "canonical_name": match_key(ingredient_name),
            "afcd_match_score": nutrition_score,
            "cpi_category": cpi_category,
            **cpi,
            "foodkeeper_product_id": expiry_match.get("foodkeeper_product_id", ""),
            "foodkeeper_name": expiry_match.get("foodkeeper_name", ""),
            "foodkeeper_subtitle": expiry_match.get("foodkeeper_subtitle", ""),
            "expiry_match_score": expiry_match.get("match_score", ""),
            "pantry_min_days": expiry.get("pantry_min_days", ""),
            "pantry_max_days": expiry.get("pantry_max_days", ""),
            "refrigerate_min_days": expiry.get("refrigerate_min_days", ""),
            "refrigerate_max_days": expiry.get("refrigerate_max_days", ""),
            "freeze_min_days": expiry.get("freeze_min_days", ""),
            "freeze_max_days": expiry.get("freeze_max_days", ""),
        }
        if nutrition:
            row.update({
                "afcd_food_key": nutrition.get("food_key", ""),
                "afcd_food_name": nutrition.get("food_name", ""),
                "nutrition_basis": nutrition.get("basis", ""),
                "energy_kj": nutrition.get("energy_kj", ""),
                "energy_kcal": nutrition.get("energy_kcal", ""),
                "protein_g": nutrition.get("protein_g", ""),
                "fat_total_g": nutrition.get("fat_total_g", ""),
                "carbohydrates_g": nutrition.get("carbohydrates_g", ""),
                "sugars_g": nutrition.get("sugars_g", ""),
                "fibre_g": nutrition.get("fibre_g", ""),
                "sodium_mg": nutrition.get("sodium_mg", ""),
            })
        rows.append(row)
    return rows


def run(args):
    output_dir = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    known = read_known_ingredients(args.known_ingredients)
    nutrition = load_afcd_nutrition(args.afcd)
    cpi = load_cpi_indexes(args.cpi)
    expiry, foodkeeper_products = load_foodkeeper_expiry(args.foodkeeper)

    expiry_aliases = build_expiry_aliases(foodkeeper_products)
    expiry_matches = []
    for ingredient in known:
        match = score_expiry_match(ingredient, expiry_aliases)
        if match:
            expiry_matches.append(match)

    latest_cpi = latest_cpi_by_category(cpi)
    food_reference = build_food_reference(known, nutrition, latest_cpi, expiry, expiry_matches)

    write_csv(output_dir / "afcd_nutrition_clean.csv", nutrition, NUTRITION_COLUMNS)
    write_csv(output_dir / "cpi_food_indexes.csv", cpi, CPI_COLUMNS)
    write_csv(output_dir / "expiry_reference.csv", expiry, EXPIRY_COLUMNS)
    write_csv(output_dir / "ingredient_expiry_matches.csv", expiry_matches, EXPIRY_MATCH_COLUMNS)
    write_csv(output_dir / "food_reference.csv", food_reference, FOOD_REFERENCE_COLUMNS)

    print(f"known_ingredients: {len(known)}")
    print(f"afcd_nutrition_clean.csv: {len(nutrition)}")
    print(f"cpi_food_indexes.csv: {len(cpi)}")
    print(f"expiry_reference.csv: {len(expiry)}")
    print(f"ingredient_expiry_matches.csv: {len(expiry_matches)}")
    print(f"food_reference.csv: {len(food_reference)}")
    print(f"output_dir: {output_dir}")


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--afcd", type=Path, default=DEFAULT_AFCD)
    parser.add_argument("--cpi", type=Path, default=DEFAULT_CPI)
    parser.add_argument("--foodkeeper", type=Path, default=DEFAULT_FOODKEEPER)
    parser.add_argument("--known-ingredients", type=Path, default=DEFAULT_KNOWN)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
